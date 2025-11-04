import os
from typing import Dict, Any, Optional, List, Tuple
import requests
import pandas as pd
import numpy as np
import streamlit as st

# ─────────────────────────────────────────────
# Safe dotenv
# ─────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# ─────────────────────────────────────────────
# Config & Keys
# ─────────────────────────────────────────────
st.set_page_config(page_title="TruLine – Weekly ML Picks + Parlays", layout="wide")
st.title("TruLine – Weekly Moneyline Picks + Parlays 🚀")
st.caption("Each sport: AI moneyline for every game + a parlay. All picks append to one Excel file with live formulas.")
st.divider()

ODDS_API_KEY = os.getenv("ODDS_API_KEY", "") or "1d677dc98d978ccc24d9914d835442f1"
DEFAULT_REGIONS = os.getenv("REGIONS", "us")  # e.g., "us,eu,uk" if you want to mix

# We only use moneyline (h2h) for this app
MARKETS = "h2h"

# Enabled sports
SOCCER_KEYS = [
    "soccer_epl",
    "soccer_spain_la_liga",
    "soccer_italy_serie_a",
    "soccer_france_ligue_one",
    "soccer_germany_bundesliga",
    "soccer_uefa_champions_league",
]

SPORT_OPTIONS: Dict[str, Any] = {
    "NFL": "americanfootball_nfl",
    "NBA": "basketball_nba",
    "MLB": "baseball_mlb",
    "NCAAF": "americanfootball_ncaaf",
    "NCAAB": "basketball_ncaab",
    "Soccer (All Major Leagues)": SOCCER_KEYS,
}

# Filter absurd lines (requested)
ODDS_MIN_US = -2000
ODDS_MAX_US = 2000

# Storage
RESULTS_XLSX = "results.xlsx"

# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
def american_to_decimal(odds: Optional[float]) -> float:
    if odds is None or pd.isna(odds):
        return np.nan
    o = float(odds)
    return 1.0 + (o / 100.0) if o > 0 else 1.0 + (100.0 / abs(o))

def implied_prob_american(odds: Optional[float]) -> float:
    if odds is None or pd.isna(odds):
        return np.nan
    o = float(odds)
    return 100.0 / (o + 100.0) if o > 0 else abs(o) / (abs(o) + 100.0)

def fmt_pct(x: float) -> str:
    return "" if (x is None or pd.isna(x)) else f"{100.0 * x:.1f}%"

def assign_units_from_score(score: float) -> float:
    score = max(0.0, min(1.0, float(score)))
    return round(0.5 + 4.5 * score, 1)

def decimal_to_american(dec: float) -> Optional[int]:
    if dec is None or pd.isna(dec) or dec <= 1.0:
        return None
    # dec odds → US odds
    if dec >= 2.0:
        return int(round((dec - 1.0) * 100))
    else:
        return int(round(-100.0 / (dec - 1.0)))

# ─────────────────────────────────────────────
# Odds API
# ─────────────────────────────────────────────
def _odds_get(url: str, params: Dict[str, Any]) -> Optional[Any]:
    if not ODDS_API_KEY:
        st.error("Missing ODDS_API_KEY.")
        return None
    try:
        r = requests.get(url, params=params, timeout=30)
        if r.status_code == 200:
            return r.json()
        else:
            return None
    except Exception:
        return None

@st.cache_data(ttl=120)
def fetch_odds(sport_key: str, regions: str, markets: str = MARKETS) -> pd.DataFrame:
    url = f"https://api.the-odds-api.com/v4/sports/{sport_key}/odds/"
    data = _odds_get(url, {
        "apiKey": ODDS_API_KEY,
        "regions": regions,
        "markets": markets,
        "oddsFormat": "american"
    })
    if not data:
        return pd.DataFrame()

    rows = []
    for ev in data:
        event_id = ev.get("id")
        commence = ev.get("commence_time")
        home = ev.get("home_team", "Unknown")
        away = ev.get("away_team", "Unknown")
        for bk in ev.get("bookmakers", []):
            book = bk.get("title")
            for mk in bk.get("markets", []):
                if mk.get("key") != "h2h":
                    continue
                for oc in mk.get("outcomes", []):
                    rows.append({
                        "sport_key": sport_key,
                        "event_id": event_id,
                        "commence_time": commence,
                        "home_team": home,
                        "away_team": away,
                        "book": book,
                        "market": "h2h",
                        "outcome": oc.get("name"),   # "Home" or "Away"
                        "line": None,                 # ML has no line
                        "odds_american": oc.get("price"),
                        "odds_decimal": american_to_decimal(oc.get("price")),
                        "conf_book": implied_prob_american(oc.get("price")),
                    })
    df = pd.DataFrame(rows)
    if df.empty:
        return df

    df["commence_time"] = pd.to_datetime(df["commence_time"], errors="coerce")
    if pd.api.types.is_datetime64tz_dtype(df["commence_time"]):
        df["Date/Time"] = df["commence_time"].dt.tz_convert("US/Eastern").dt.strftime("%b %d, %I:%M %p ET")
    else:
        df["Date/Time"] = df["commence_time"].dt.tz_localize("UTC").dt.tz_convert("US/Eastern").dt.strftime("%b %d, %I:%M %p ET")
    return df

def fetch_sport(label: str, s) -> pd.DataFrame:
    if isinstance(s, list):
        frames = []
        for sub in s:
            d = fetch_odds(sub, st.session_state.regions)
            if not d.empty:
                d["sport_label"] = label
                frames.append(d)
        raw = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    else:
        raw = fetch_odds(s, st.session_state.regions)
        if not raw.empty:
            raw["sport_label"] = label

    if raw.empty:
        return raw

    # ✅ filter to next 7 days only
    now = pd.Timestamp.utcnow()
    future_7 = now + pd.Timedelta(days=7)
    raw = raw[(raw["commence_time"] >= now) & (raw["commence_time"] <= future_7)]

    return raw

# ─────────────────────────────────────────────
# Consensus + Ensemble (simple ML only)
# ─────────────────────────────────────────────
def build_consensus_ml(raw: pd.DataFrame) -> pd.DataFrame:
    if raw.empty:
        return raw

    # Best price per (event, outcome)
    idx_best = raw.groupby(["event_id", "market", "outcome"])["odds_decimal"].idxmax()
    best = raw.loc[idx_best, ["event_id","market","outcome","odds_american","odds_decimal","book"]]
    best = best.rename(columns={
        "odds_american":"best_odds_us",
        "odds_decimal":"best_odds_dec",
        "book":"best_book"
    })

    agg = raw.groupby(["event_id","market","outcome"], dropna=False).agg(
        consensus_conf=("conf_book","mean"),
        books=("book","nunique"),
        home_team=("home_team","first"),
        away_team=("away_team","first"),
        commence_time=("commence_time","first"),
        date_time=("Date/Time","first"),
        sport=("sport_label","first"),
        avg_odds_dec=("odds_decimal","mean"),
    ).reset_index()

    out = agg.merge(best, on=["event_id","market","outcome"], how="left")
    out["Matchup"] = out["home_team"] + " vs " + out["away_team"]
    out["Date/Time"] = out["date_time"]

    # Filter insane odds
    out = out[(out["best_odds_us"] >= ODDS_MIN_US) & (out["best_odds_us"] <= ODDS_MAX_US)]

    return out[[
        "sport","event_id","commence_time","Date/Time","Matchup","market","outcome",
        "best_book","best_odds_us","best_odds_dec","consensus_conf","books","avg_odds_dec"
    ]]

def ensemble_score_ml(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    d = df.copy()

    # Voters
    v1 = d["consensus_conf"].astype(float).fillna(0.0).clip(0.0, 1.0)
    edge = (d["best_odds_dec"] - d["avg_odds_dec"]).astype(float).fillna(0.0)
    v2 = 1.0 / (1.0 + np.exp(-6.0 * edge))
    v3 = (d["books"].astype(float).clip(0.0, 10.0)) / 10.0
    dec = d["best_odds_dec"].astype(float).fillna(2.0)
    v4 = np.exp(-((dec - 2.0) ** 2) / (2 * (0.6 ** 2)))
    v4 = (v4 - v4.min()) / (v4.max() - v4.min() + 1e-9)

    d["V1_prob"] = v1
    d["V2_edge"] = v2
    d["V3_depth"] = v3
    d["V4_balance"] = v4
    d["EnsembleScore"] = (v1 + v2 + v3 + v4) / 4.0
    d["Units"] = d["EnsembleScore"].apply(assign_units_from_score)
    return d

def pick_side_for_game(scored: pd.DataFrame) -> pd.DataFrame:
    """
    From (event, outcome in {Home, Away}), keep the better score as the pick.
    """
    if scored.empty:
        return pd.DataFrame()
    idx = scored.groupby(["event_id"])["EnsembleScore"].idxmax()
    best = scored.loc[idx].copy()
    best["Market"] = "Moneyline"
    best["Pick"] = best["outcome"]
    best["Odds (US)"] = best["best_odds_us"].astype(int)
    best["Odds (Dec)"] = best["best_odds_dec"].round(3)
    best["Confidence"] = best["consensus_conf"].apply(fmt_pct)
    best = best.rename(columns={"sport": "Sport", "best_book": "Sportsbook"})
    return best[[
        "Date/Time","Sport","Matchup","Market","Pick","Sportsbook","Odds (US)","Odds (Dec)",
        "Confidence","Units","EnsembleScore","event_id"
    ]].sort_values("Date/Time").reset_index(drop=True)

def build_moneyline_parlay(picks: pd.DataFrame, max_legs: int = 8) -> Optional[Dict[str, Any]]:
    if picks is None or picks.empty:
        return None
    top = picks.copy().sort_values("EnsembleScore", ascending=False).head(max_legs)
    if top.empty:
        return None
    dec_odds = 1.0
    units = 0.0
    for _, r in top.iterrows():
        d = float(r["Odds (Dec)"])
        if d <= 1.0 or pd.isna(d):
            d = 1.01
        dec_odds *= d
        units += float(r["Units"])
    return {
        "legs": len(top),
        "parlay_decimal": round(dec_odds, 4),
        "parlay_american": decimal_to_american(dec_odds),
        "suggested_units": round(max(0.5, min(5.0, units * 0.4)), 1),
        "legs_df": top.reset_index(drop=True)
    }

# ─────────────────────────────────────────────
# Excel Append (one sheet, formulas at top)
# ─────────────────────────────────────────────
RESULTS_COLUMNS = [
    "Date/Time","Sport","Matchup","Market","Pick","Odds_US","Odds_Dec","Units","Result"
]

def ensure_results_headers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure dataframe has Excel output columns in correct order.
    """
    out = pd.DataFrame(columns=RESULTS_COLUMNS)
    if df is None or df.empty:
        return out
    tmp = df.rename(columns={
        "Odds (US)":"Odds_US",
        "Odds (Dec)":"Odds_Dec"
    }).copy()
    for col in RESULTS_COLUMNS:
        if col not in tmp.columns:
            tmp[col] = ""
    tmp = tmp[RESULTS_COLUMNS].copy()
    tmp["Units"] = pd.to_numeric(tmp["Units"], errors="coerce").fillna(1.0)
    tmp["Result"] = tmp["Result"].fillna("") if "Result" in tmp.columns else ""
    return tmp

def read_results_excel() -> pd.DataFrame:
    if not os.path.exists(RESULTS_XLSX):
        return pd.DataFrame(columns=RESULTS_COLUMNS)
    try:
        x = pd.read_excel(RESULTS_XLSX, sheet_name="Results")
        # drop the first two rows if they are the summary header/formula rows
        # Heuristic: first two rows do not match the columns; or contain "SUMMARY"
        if len(x) >= 1 and ("SUMMARY" in str(x.iloc[0,0]).upper() or "LABEL" in str(x.iloc[0,0]).upper()):
            # This happens if someone edited the sheet incorrectly—fallback: reload with header=2
            x = pd.read_excel(RESULTS_XLSX, sheet_name="Results", header=2)
        x = x.rename(columns={"Odds (US)":"Odds_US","Odds (Dec)":"Odds_Dec"})
        for col in RESULTS_COLUMNS:
            if col not in x.columns:
                x[col] = "" if col not in ["Units"] else 1.0
        return x[RESULTS_COLUMNS].copy()
    except Exception:
        return pd.DataFrame(columns=RESULTS_COLUMNS)

def write_results_excel(all_rows: pd.DataFrame):
    """
    Writes the entire Results sheet and places summary formulas in row 1/2.
    """
    # We will write with a blank first two rows reserved for formulas.
    # Strategy: write data starting at row=3 (header in row=3)
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Create workbook or load existing
    if os.path.exists(RESULTS_XLSX):
        try:
            wb = load_workbook(RESULTS_XLSX)
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()

    # Remove existing "Results" sheet if exists
    if "Results" in wb.sheetnames:
        ws = wb["Results"]
        wb.remove(ws)
    ws = wb.create_sheet("Results")

    # Row 1: labels + formulas across columns B..G
    ws["A1"].value = "SUMMARY"
    ws["B1"].value = "Total"
    ws["C1"].value = "Wins"
    ws["D1"].value = "Losses"
    ws["E1"].value = "Win %"
    ws["F1"].value = "Units Won"
    ws["G1"].value = "ROI"

    # Row 2: formulas using full columns (Result is I, Units is H)
    ws["B2"].value = '=COUNTA(A:A)-2'
    ws["C2"].value = '=COUNTIF(I:I,"Win")'
    ws["D2"].value = '=COUNTIF(I:I,"Loss")'
    ws["E2"].value = '=IFERROR(C2/(C2+D2),0)'
    ws["F2"].value = '=(SUMPRODUCT((I3:I="Win")*(H3:H)) - SUMPRODUCT((I3:I="Loss")*(H3:H)))'
    ws["G2"].value = '=IFERROR(F2 / SUMPRODUCT(((I3:I="Win")+(I3:I="Loss"))*(H3:H)),0)'

    # Leave row 2 otherwise blank; write header at row=3, then data from row=4
    # Write header
    for j, col in enumerate(RESULTS_COLUMNS, start=1):
        ws.cell(row=3, column=j).value = col

    # Write data rows
    if not all_rows.empty:
        for r in dataframe_to_rows(all_rows, index=False, header=False):
            ws.append(r)

    # If workbook had default "Sheet", remove it
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])

    wb.save(RESULTS_XLSX)

def append_picks_to_excel(picks_df: pd.DataFrame):
    """
    Read current Results, append deduped, and write with formulas.
    """
    current = read_results_excel()
    to_add = ensure_results_headers(picks_df)
    if to_add.empty:
        # nothing to add → still ensure the file has formulas
        write_results_excel(current)
        return

    all_rows = pd.concat([current, to_add], ignore_index=True)
    # Dedup on identity keys (Date/Time,Sport,Matchup,Market,Pick)
    all_rows = all_rows.drop_duplicates(subset=["Date/Time","Sport","Matchup","Market","Pick"], keep="first")
    # Keep chronological
    all_rows = all_rows.copy()
    # Date/Time is a string; leave as is to avoid TZ issues
    write_results_excel(all_rows)

# ─────────────────────────────────────────────
# Sidebar controls
# ─────────────────────────────────────────────
with st.sidebar:
    st.subheader("Controls")
    st.session_state.regions = st.text_input("Regions", value=DEFAULT_REGIONS)
    # Parlay legs per sport (min 3, max 10-ish; you asked ~3–10)
    parlay_legs = st.slider("Parlay legs per sport", 3, 10, 5)
    generate = st.button("Generate Weekly Picks")

# ─────────────────────────────────────────────
# Generate per-sport ML picks + parlays; append to Excel
# ─────────────────────────────────────────────
if generate:
    st.session_state.all_sport_picks = {}
    st.session_state.all_sport_parlays = {}
    st.session_state.all_for_excel = []

    for label, key in SPORT_OPTIONS.items():
        raw = fetch_sport(label, key)
        if raw.empty:
            st.session_state.all_sport_picks[label] = pd.DataFrame()
            st.session_state.all_sport_parlays[label] = None
            continue

        cons = build_consensus_ml(raw)
        if cons.empty:
            st.session_state.all_sport_picks[label] = pd.DataFrame()
            st.session_state.all_sport_parlays[label] = None
            continue

        scored = ensemble_score_ml(cons)
        picks = pick_side_for_game(scored)  # ML pick for every game in that sport
        st.session_state.all_sport_picks[label] = picks

        parlay = build_moneyline_parlay(picks, max_legs=parlay_legs)
        st.session_state.all_sport_parlays[label] = parlay

        if not picks.empty:
            # Prepare for Excel append
            x = picks.rename(columns={"Odds (US)":"Odds_US","Odds (Dec)":"Odds_Dec"}).copy()
            x["Market"] = "Moneyline"
            x["Result"] = ""  # user will mark Win/Loss in Excel later
            x = x[["Date/Time","Sport","Matchup","Market","Pick","Odds_US","Odds_Dec","Units","Result"]]
            st.session_state.all_for_excel.append(x)

    # Append all picks from all sports to Excel (one sheet)
    if st.session_state.all_for_excel:
        all_new_rows = pd.concat(st.session_state.all_for_excel, ignore_index=True)
        append_picks_to_excel(all_new_rows)
    else:
        # Still ensure file exists and has formulas
        append_picks_to_excel(pd.DataFrame(columns=RESULTS_COLUMNS))

    st.session_state.has_data = True

# ─────────────────────────────────────────────
# Render tabs per sport
# ─────────────────────────────────────────────
if st.session_state.get("has_data", False):
    tabs = st.tabs(list(SPORT_OPTIONS.keys()) + ["📥 Export"])

    # Per sport tabs
    for idx, label in enumerate(SPORT_OPTIONS.keys()):
        with tabs[idx]:
            st.subheader(f"{label} — Moneyline Picks (this week)")
            picks = st.session_state.all_sport_picks.get(label, pd.DataFrame())
            if picks is None or picks.empty:
                st.info("No picks found.")
            else:
                show = picks.copy()
                show = show.rename(columns={"EnsembleScore":"Score"})
                show["Score"] = show["Score"].map(lambda x: f"{float(x):.3f}")
                st.dataframe(show.drop(columns=["event_id"]), use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader(f"{label} — Suggested Parlay")
            parlay = st.session_state.all_sport_parlays.get(label)
            if parlay is None:
                st.info("No parlay could be built.")
            else:
                c1,c2,c3 = st.columns(3)
                c1.metric("Legs", str(parlay["legs"]))
                c2.metric("Parlay Odds (Dec)", str(parlay["parlay_decimal"]))
                c3.metric("Suggested Units", str(parlay["suggested_units"]))
                st.caption(f"Parlay Odds (US): {parlay['parlay_american'] if parlay['parlay_american'] is not None else 'N/A'}")

                with st.expander("Show Parlay Legs"):
                    legs_df = parlay["legs_df"].rename(columns={"EnsembleScore":"Score"}).copy()
                    legs_df["Score"] = legs_df["Score"].map(lambda x: f"{float(x):.3f}")
                    st.dataframe(legs_df.drop(columns=["event_id"]), use_container_width=True, hide_index=True)
# ──────────────────────────────────
# Manual Result Updater
# ──────────────────────────────────
with tabs.insert(len(SPORT_OPTIONS), "Update Results"):
    st.subheader("Update Results (Win / Loss)")

    results_df = read_results_excel().copy()

    if results_df.empty:
        st.info("No picks saved yet.")
    else:
        for i in range(len(results_df)):
            c1, c2, c3 = st.columns([6,2,1])
            with c1:
                st.write(f"{results_df.iloc[i]['Date/Time']} — {results_df.iloc[i]['Sport']} — {results_df.iloc[i]['Matchup']}")
            with c2:
                new_val = st.selectbox(
                    "",
                    ["", "Win", "Loss"],
                    index=["","Win","Loss"].index(str(results_df.iloc[i]["Result"])),
                    key=f"result_row_{i}"
                )
                results_df.at[i,"Result"] = new_val

        if st.button("Save Result Updates"):
            write_results_excel(results_df)
            st.success("✅ Results updated")
            st.experimental_rerun()
    # Export tab
    with tabs[-1]:
        st.subheader("Results Excel (one sheet)")
        st.write("• All picks append to **results.xlsx → Results**")
        st.write("• Mark the **Result** column (Win/Loss) in Excel; summary formulas at the top auto-update.")
        # Offer download of current file contents (if exists)
        if os.path.exists(RESULTS_XLSX):
            with open(RESULTS_XLSX, "rb") as f:
                data = f.read()
            st.download_button(
                "Download results.xlsx",
                data=data,
                file_name="results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("No results.xlsx yet. Generate picks first.")
else:
    st.info("Click **Generate Weekly Picks** in the sidebar.")
